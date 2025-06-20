from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

def create_monthly_sheet(wb, df, jaar, suffix=""):
    """Maandelijks overzicht van inkomsten en uitgaven"""
    ws = wb.create_sheet(f"Maandoverzicht{suffix}")

    # Zet maand als periode (bijv. 2024-03)
    df_monthly = df.copy()
    df_monthly['maand'] = pd.to_datetime(df_monthly['datum']).dt.to_period('M')

    # Groepeer per maand
    inkomsten = df_monthly[df_monthly['bedrag'] > 0].groupby('maand')['bedrag'].sum()
    uitgaven = df_monthly[df_monthly['bedrag'] < 0].groupby('maand')['bedrag'].sum().abs()

    # Maak volledige lijst van maanden van januari t/m december
    alle_maanden = pd.period_range(start=f"{jaar}-01", end=f"{jaar}-12", freq='M')

    # Combineer in DataFrame
    monthly_summary = pd.DataFrame(index=alle_maanden)
    monthly_summary['inkomsten'] = inkomsten
    monthly_summary['uitgaven'] = uitgaven
    monthly_summary = monthly_summary.fillna(0)
    monthly_summary['netto'] = monthly_summary['inkomsten'] - monthly_summary['uitgaven']

    # Titel en headers
    ws['A1'] = f"Maandoverzicht {jaar}{suffix}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    headers = ['Maand', 'Inkomsten', 'Uitgaven', 'Netto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Data invullen
    for i, (maand, row) in enumerate(monthly_summary.iterrows(), start=4):
        ws.cell(row=i, column=1, value=str(maand))
        ws.cell(row=i, column=2, value=row['inkomsten']).number_format = '#,##0.00'
        ws.cell(row=i, column=3, value=row['uitgaven']).number_format = '#,##0.00'
        ws.cell(row=i, column=4, value=row['netto']).number_format = '#,##0.00'

    # Auto-breedte
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
