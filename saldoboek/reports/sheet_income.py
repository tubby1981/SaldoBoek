def create_income_sheet(wb, df, jaar, suffix=""):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(f"Inkomsten{suffix}")
    inkomsten_df = df[df['bedrag'] > 0].copy()

    if inkomsten_df.empty:
        ws['A1'] = "Geen inkomsten gevonden"
        return

    categorie_totalen = inkomsten_df.groupby('categorie')['bedrag'].agg(['sum', 'count']).reset_index()
    categorie_totalen.columns = ['Categorie', 'Totaal', 'Aantal transacties']
    categorie_totalen = categorie_totalen.sort_values('Totaal', ascending=False)

    ws['A1'] = f"Inkomsten {jaar}{suffix}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.append(["Categorie", "Totaal bedrag", "Aantal transacties", "Gemiddeld per transactie"])
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    for _, row in categorie_totalen.iterrows():
        ws.append([
            row['Categorie'],
            row['Totaal'],
            row['Aantal transacties'],
            row['Totaal'] / row['Aantal transacties']
        ])

    totaal = categorie_totalen['Totaal'].sum()
    totaal_excl = categorie_totalen[~categorie_totalen['Categorie'].isin(['Overboekingen ontvangen'])]['Totaal'].sum()
    aantal = categorie_totalen['Aantal transacties'].sum()
    aantal_excl = categorie_totalen[~categorie_totalen['Categorie'].isin(['Overboekingen ontvangen'])]['Aantal transacties'].sum()

    ws.append(["TOTAAL (excl. Overboekingen ontvangen)", totaal_excl, aantal_excl, ""])
    ws.append(["TOTAAL (incl. alles)", totaal, aantal, ""])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
