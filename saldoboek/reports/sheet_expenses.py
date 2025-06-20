def create_expenses_sheet(wb, df, jaar, suffix=""):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(f"Uitgaven{suffix}")
    uitgaven_df = df[df['bedrag'] < 0].copy()
    uitgaven_df['bedrag'] = abs(uitgaven_df['bedrag'])

    if uitgaven_df.empty:
        ws['A1'] = "Geen uitgaven gevonden"
        return

    categorie_totalen = uitgaven_df.groupby('categorie')['bedrag'].agg(['sum', 'count']).reset_index()
    categorie_totalen.columns = ['Categorie', 'Totaal', 'Aantal transacties']
    categorie_totalen = categorie_totalen.sort_values('Totaal', ascending=False)

    ws['A1'] = f"Uitgaven {jaar}{suffix}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.append(["Categorie", "Totaal bedrag", "Aantal transacties", "Gemiddeld per transactie", "% van totaal"])
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    totaal_uitgaven = categorie_totalen['Totaal'].sum()
    for _, row in categorie_totalen.iterrows():
        ws.append([
            row['Categorie'],
            row['Totaal'],
            row['Aantal transacties'],
            row['Totaal'] / row['Aantal transacties'],
            row['Totaal'] / totaal_uitgaven
        ])

    ws.append(["TOTAAL", totaal_uitgaven, categorie_totalen['Aantal transacties'].sum(), "", 1.0])
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
