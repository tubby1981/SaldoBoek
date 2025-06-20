from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

def create_monthly_category_sheet(wb, df, jaar, suffix=""):
    ws = wb.create_sheet(f"Maand-Categorie{suffix}")

    df_analysis = df.copy()
    df_analysis['maand'] = pd.to_datetime(df_analysis['datum']).dt.to_period('M')

    inkomsten_df = df_analysis[df_analysis['bedrag'] > 0]
    uitgaven_df = df_analysis[df_analysis['bedrag'] < 0].copy()
    uitgaven_df['bedrag'] = abs(uitgaven_df['bedrag'])

    ws['A1'] = f"Maand-Categorie Analyse {jaar}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')

    current_row = 3

    ws[f'A{current_row}'] = "INKOMSTEN PER MAAND EN CATEGORIE"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="008000")
    current_row += 2

    if not inkomsten_df.empty:
        inkomsten_pivot = inkomsten_df.pivot_table(
            index='categorie',
            columns='maand',
            values='bedrag',
            aggfunc='sum',
            fill_value=0
        )

        ws[f'A{current_row}'] = "Categorie"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

        col = 2
        for maand in inkomsten_pivot.columns:
            cell = ws.cell(row=current_row, column=col, value=str(maand))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            col += 1

        ws.cell(row=current_row, column=col, value="Totaal").font = Font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

        current_row += 1

        for categorie in inkomsten_pivot.index:
            ws[f'A{current_row}'] = categorie
            col = 2
            row_total = 0
            for maand in inkomsten_pivot.columns:
                bedrag = inkomsten_pivot.loc[categorie, maand]
                if bedrag > 0:
                    cell = ws.cell(row=current_row, column=col, value=bedrag)
                    cell.number_format = '#,##0.00'
                    row_total += bedrag
                col += 1
            if row_total > 0:
                cell = ws.cell(row=current_row, column=col, value=row_total)
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)
            current_row += 1

        ws[f'A{current_row}'] = "TOTAAL INKOMSTEN"
        ws[f'A{current_row}'].font = Font(bold=True)
        col = 2
        grand_total = 0
        for maand in inkomsten_pivot.columns:
            maand_total = inkomsten_pivot[maand].sum()
            if maand_total > 0:
                cell = ws.cell(row=current_row, column=col, value=maand_total)
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)
                grand_total += maand_total
            col += 1
        cell = ws.cell(row=current_row, column=col, value=grand_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, color="008000")

        current_row += 3

    ws[f'A{current_row}'] = "UITGAVEN PER MAAND EN CATEGORIE"
    ws[f'A{current_row}'].font = Font(bold=True, size=14, color="CC0000")
    current_row += 2

    if not uitgaven_df.empty:
        uitgaven_pivot = uitgaven_df.pivot_table(
            index='categorie',
            columns='maand',
            values='bedrag',
            aggfunc='sum',
            fill_value=0
        )

        ws[f'A{current_row}'] = "Categorie"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        col = 2
        for maand in uitgaven_pivot.columns:
            cell = ws.cell(row=current_row, column=col, value=str(maand))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            col += 1

        ws.cell(row=current_row, column=col, value="Totaal").font = Font(bold=True)
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        current_row += 1

        for categorie in uitgaven_pivot.index:
            ws[f'A{current_row}'] = categorie
            col = 2
            row_total = 0
            for maand in uitgaven_pivot.columns:
                bedrag = uitgaven_pivot.loc[categorie, maand]
                if bedrag > 0:
                    cell = ws.cell(row=current_row, column=col, value=bedrag)
                    cell.number_format = '#,##0.00'
                    row_total += bedrag
                col += 1
            if row_total > 0:
                cell = ws.cell(row=current_row, column=col, value=row_total)
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)
            current_row += 1

        ws[f'A{current_row}'] = "TOTAAL UITGAVEN"
        ws[f'A{current_row}'].font = Font(bold=True)
        col = 2
        grand_total = 0
        for maand in uitgaven_pivot.columns:
            maand_total = uitgaven_pivot[maand].sum()
            if maand_total > 0:
                cell = ws.cell(row=current_row, column=col, value=maand_total)
                cell.number_format = '#,##0.00'
                cell.font = Font(bold=True)
                grand_total += maand_total
            col += 1
        cell = ws.cell(row=current_row, column=col, value=grand_total)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True, color="CC0000")

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)
