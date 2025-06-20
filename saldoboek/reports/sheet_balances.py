from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

def create_balance_sheet(wb, df, jaar, suffix=""):
    ws = wb.create_sheet(f"Rekening Saldi{suffix}")

    ws['A1'] = f"Rekening Overzicht {jaar}"
    ws['A1'].font = Font(bold=True, size=14)

    headers = [
        "Rekening", "Type", "Aantal transacties", "Totale mutatie",
        "Eerste transactie", "Laatste transactie",
        "Beginsaldo (1 jan)", "Eindsaldo (31 dec)"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

    row = 4
    for rekening in df['rekening'].unique():
        df_rek = df[df['rekening'] == rekening].sort_values('datum')
        if df_rek.empty:
            continue

        eerste = df_rek.iloc[0]
        laatste = df_rek.iloc[-1]

        beginsaldo = eerste['saldo_voor']
        eindsaldo = laatste['saldo_voor'] + laatste['bedrag']

        try:
            eerste_datum = pd.to_datetime(eerste['datum']).strftime('%Y-%m-%d')
            laatste_datum = pd.to_datetime(laatste['datum']).strftime('%Y-%m-%d')
        except:
            eerste_datum = str(eerste['datum'])
            laatste_datum = str(laatste['datum'])

        ws.cell(row=row, column=1, value=rekening)
        ws.cell(row=row, column=2, value=eerste['rekeningtype'])
        ws.cell(row=row, column=3, value=len(df_rek))
        ws.cell(row=row, column=4, value=df_rek['bedrag'].sum()).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=eerste_datum)
        ws.cell(row=row, column=6, value=laatste_datum)
        ws.cell(row=row, column=7, value=beginsaldo).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=eindsaldo).number_format = '#,##0.00'

        row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
