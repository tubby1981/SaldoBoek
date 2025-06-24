import pandas as pd

def get_interne_overboekingen_per_rekening(df):
    sparen = df[(df['bedrag'] < 0) & (df['categorie'] == 'Sparen/Overboeken')].copy()
    sparen['bedrag'] = sparen['bedrag'].abs()
    ob_ontvangen = df[(df['bedrag'] > 0) & (df['categorie'] == 'Overboekingen ontvangen')]
    rente = df[(df['bedrag'] > 0) & (df['categorie'] == 'Rente')]
    
    # Sparen/overboeken per rekening
    sparen_per_rekening = sparen.groupby('rekening')['bedrag'].sum().reset_index(name='sparen')
    ob_per_rekening = ob_ontvangen.groupby('rekening')['bedrag'].sum().reset_index(name='ontvangen')
    rente_per_rekening = rente.groupby('rekening')['bedrag'].sum().reset_index(name='rente')
    
    from functools import reduce
    dfs = [sparen_per_rekening, ob_per_rekening, rente_per_rekening]
    samengevoegd = reduce(lambda left, right: pd.merge(left, right, on='rekening', how='outer'), dfs)
    samengevoegd = samengevoegd.fillna(0)
    return samengevoegd

def create_overview_sheet(wb, df, jaar, suffix=""):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    ws = wb.create_sheet(f"Overzicht{suffix}")
    ws['A1'] = f"Jaaroverzicht {jaar}{suffix}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A3'] = "Rapportage periode:"
    ws['B3'] = f"01-01-{jaar} t/m 31-12-{jaar}"
    ws['A4'] = "Gegenereerd op:"
    ws['B4'] = datetime.now().strftime("%d-%m-%Y %H:%M")
    
    # Filter dataframes
    inkomsten_df = df[df['bedrag'] > 0].copy()
    uitgaven_df = df[df['bedrag'] < 0].copy()
    uitgaven_df['bedrag'] = uitgaven_df['bedrag'].abs()
    
    # Bereken externe bedragen (zonder interne overboekingen)
    externe_inkomsten = inkomsten_df[(inkomsten_df['categorie'] != 'Overboekingen ontvangen')]['bedrag'].sum()
    externe_uitgaven = uitgaven_df[~uitgaven_df['categorie'].isin(['Sparen/Overboeken'])]['bedrag'].sum()
    extern_netto = externe_inkomsten - externe_uitgaven
    
    # Bereken totale bedragen (inclusief interne overboekingen)
    totale_inkomsten = inkomsten_df['bedrag'].sum()
    totale_uitgaven = uitgaven_df['bedrag'].sum()
    totaal_netto = totale_inkomsten - totale_uitgaven
    
    ws.append(["", ""])
    ws.append(["FINANCIEEL OVERZICHT (netto)"])
    ws.append(["Netto inkomsten:", externe_inkomsten])
    ws.append(["Netto uitgaven:", externe_uitgaven])
    ws.append(["Netto resultaat:", extern_netto])
    
    ws.append(["", ""])
    ws.append(["FINANCIEEL OVERZICHT (totaal inclusief interne overboekingen)"])
    ws.append(["Totale inkomsten:", totale_inkomsten])
    ws.append(["Totale uitgaven:", totale_uitgaven])
    ws.append(["Totaal netto resultaat:", totaal_netto])
    
    # Interne overboekingen per rekening
    ws.append(["", ""])
    ws.append(["INTERNE OVERBOEKINGEN (per rekening)"])
    interne = get_interne_overboekingen_per_rekening(df)
    ws.append(["Rekening", "Ontvangen", "Sparen/Overboeken", "Rente"])
    for _, row in interne.iterrows():
        ws.append([
            row['rekening'],
            row['ontvangen'],
            row['sparen'],
            row['rente']
        ])
    ws.append([
        "Totaal",
        interne['ontvangen'].sum(),
        interne['sparen'].sum(),
        interne['rente'].sum()
    ])
    
    # Algemene statistieken
    ws.append(["", ""])
    ws.append(["STATISTIEKEN"])
    ws.append(["Aantal transacties:", len(df)])
    ws.append(["Aantal rekeningen:", df['rekening'].nunique()])
    ws.append(["Aantal categorieën:", df['categorie'].nunique()])
    
    # Kolombreedte optimaliseren
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)
