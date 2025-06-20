from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def create_transactions_sheet(wb, df, jaar, suffix=""):
    ws = wb.create_sheet(f"Alle Transacties {suffix}")

    headers = ['Datum', 'Rekening', 'Naam', 'Omschrijving', 'Bedrag', 'Categorie', 'Type']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    for row, (_, transaction) in enumerate(df.iterrows(), 2):
        ws.cell(row=row, column=1, value=transaction['datum'])
        ws.cell(row=row, column=2, value=transaction['rekening'])
        ws.cell(row=row, column=3, value=transaction['naam'])
        ws.cell(row=row, column=4, value=transaction['omschrijving'])

        bedrag_cell = ws.cell(row=row, column=5, value=transaction['bedrag'])
        bedrag_cell.number_format = '#,##0.00'
        if transaction['bedrag'] < 0:
            bedrag_cell.font = Font(color="FF0000")

        ws.cell(row=row, column=6, value=transaction['categorie'])
        ws.cell(row=row, column=7, value=transaction['rekeningtype'])

    for column in ws.columns:
        max_length = 0
        column_index = column[0].column
        column_letter = get_column_letter(column_index)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
